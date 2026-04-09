"""
Energiatodistusrekisteri - Web-sovellus
Käynnistä: python app.py
Avaa selain: http://localhost:5000
"""

from flask import Flask, request, jsonify, render_template_string
import pandas as pd
import json
import os
import re
import sys
import urllib.request
import urllib.parse
import html as html_lib
import zipfile
import io
import sqlite3
import threading

# Lataa .env tiedostosta jos löytyy (ennen muuta koodia)
_env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
if os.path.exists(_env_path):
    with open(_env_path, encoding='utf-8') as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and not _line.startswith('#') and '=' in _line:
                _k, _, _v = _line.partition('=')
                os.environ[_k.strip()] = _v.strip()  # .env ylikirjoittaa aina

app = Flask(__name__)

# ──────────────────────────────────────────────
# SQLite-tietokanta hyväksytyille yhtiönimille
# ──────────────────────────────────────────────
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'yhtiot.db')
db_lock = threading.Lock()

def init_db():
    with sqlite3.connect(DB_PATH) as con:
        con.execute('''CREATE TABLE IF NOT EXISTS yhtiot (
            rakennustunnus TEXT PRIMARY KEY,
            osoite         TEXT,
            yhtio_nimi     TEXT NOT NULL,
            omistaja       TEXT,
            ytunnus        TEXT,
            hyvaksytty_at  TEXT DEFAULT (datetime('now','localtime'))
        )''')
        # Lisää puuttuvat sarakkeet vanhoihin kantoihin
        for col_def in [
            'ALTER TABLE yhtiot ADD COLUMN omistaja TEXT',
            'ALTER TABLE yhtiot ADD COLUMN kontaktoitu_at TEXT',
            'ALTER TABLE yhtiot ADD COLUMN muistiinpanot TEXT',
        ]:
            try:
                con.execute(col_def)
            except Exception:
                pass
        con.commit()
    print("✅  Tietokanta valmis.\n")

def hae_tallennetut():
    """Palauttaa kaikki hyväksytyt yhtiönimet dict-muodossa {rakennustunnus: {...}}"""
    with sqlite3.connect(DB_PATH) as con:
        con.row_factory = sqlite3.Row
        rows = con.execute('SELECT * FROM yhtiot').fetchall()
    return {r['rakennustunnus']: dict(r) for r in rows}

init_db()

# ──────────────────────────────────────────────
# Datan lataus
# ──────────────────────────────────────────────
BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
CSV_PATH  = os.path.join(BASE_DIR, 'ethaku.csv')
ZIP_PATH  = os.path.join(BASE_DIR, 'ethaku.csv.zip')

# Pura zip automaattisesti jos CSV puuttuu mutta zip löytyy
if not os.path.exists(CSV_PATH) and os.path.exists(ZIP_PATH):
    print("⏳  Puretaan ethaku.csv.zip...")
    with zipfile.ZipFile(ZIP_PATH, 'r') as zf:
        csv_members = [m for m in zf.namelist() if m.lower().endswith('.csv')]
        print(f"   Zip sisältää: {zf.namelist()}")
        if csv_members:
            # Pura CSV suoraan BASE_DIR:iin nimellä ethaku.csv
            with zf.open(csv_members[0]) as src, open(CSV_PATH, 'wb') as dst:
                dst.write(src.read())
            print(f"✅  Purettu ({csv_members[0]} → ethaku.csv)\n")
        else:
            print("❌  Zip ei sisällä CSV-tiedostoa!")
            sys.exit(1)

if not os.path.exists(CSV_PATH):
    print(f"\n❌  Tiedostoa '{CSV_PATH}' ei löydy.")
    print("   Kopioi ethaku.csv tai ethaku.csv.zip samaan kansioon kuin app.py.\n")
    sys.exit(1)

print("⏳  Ladataan CSV-data (voi kestää hetken)...")
df = pd.read_csv(CSV_PATH, sep=';', low_memory=False)
df['Lahtotiedot / Lammitetty-nettoala'] = pd.to_numeric(
    df['Lahtotiedot / Lammitetty-nettoala'], errors='coerce'
)
print(f"✅  Ladattu {len(df):,} energiatodistusta.\n")

# ──────────────────────────────────────────────
# Postinumero → kaupunki -kartoitus (GeoNames)
# ──────────────────────────────────────────────
POSTI_CACHE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'postinumerot.json')

def lataa_postinumerot():
    """Lataa postinumerot Postin BAF-tiedostosta tai välimuistista.

    Päivitä data: lataa uusi BAF_YYYYMMDD.zip osoitteesta https://www.posti.fi/webpcode/
    ja pura BAF_YYYYMMDD.dat samaan kansioon kuin app.py. Käynnistä sovellus uudelleen.
    """
    # Tarkista löytyykö BAF-tiedosto (Postin virallinen data)
    baf_tiedostot = sorted([
        f for f in os.listdir(os.path.dirname(os.path.abspath(__file__)))
        if f.startswith('BAF_') and f.endswith('.dat')
    ], reverse=True)

    if baf_tiedostot:
        baf_polku = os.path.join(os.path.dirname(os.path.abspath(__file__)), baf_tiedostot[0])
        print(f"⏳  Luetaan postinumerodata Postin BAF-tiedostosta ({baf_tiedostot[0]})...")
        try:
            mapping = {}
            with open(baf_polku, encoding='latin-1') as f:
                for line in f:
                    if len(line) < 236:
                        continue
                    postinumero = line[13:18].strip()
                    kuntanimi   = line[216:236].strip().title()
                    if postinumero and kuntanimi:
                        mapping[postinumero] = kuntanimi
            with open(POSTI_CACHE, 'w', encoding='utf-8') as f:
                json.dump(mapping, f, ensure_ascii=False)
            print(f"✅  Luettu {len(mapping):,} postinumeroa, {len(set(mapping.values())):,} kuntaa.\n")
            return mapping
        except Exception as e:
            print(f"⚠️  BAF-tiedoston luku epäonnistui ({e}).\n")

    # Käytetään välimuistia jos BAF ei löydy
    if os.path.exists(POSTI_CACHE):
        with open(POSTI_CACHE, encoding='utf-8') as f:
            data = json.load(f)
        print(f"✅  Ladattu {len(data):,} postinumeroa välimuistista.\n")
        return data

    print("⚠️  Postinumerodataa ei löydy. Kaupunkisuodatin ei ole käytössä.\n")
    return {}

POSTI_KAUPUNKI = lataa_postinumerot()

# Rakennetaan kaupunki → [postinumerot] -hakemisto
KAUPUNKI_POSTIT = {}
for koodi, kaupunki in POSTI_KAUPUNKI.items():
    KAUPUNKI_POSTIT.setdefault(kaupunki, []).append(koodi)
KAUPUNGIT = sorted(KAUPUNKI_POSTIT.keys())

# ──────────────────────────────────────────────
# Normalisoidut ryhmäsarakkeet
# ──────────────────────────────────────────────

def normalisoi_lammitys(arvo):
    if pd.isna(arvo):
        return 'Muu / tuntematon'
    a = str(arvo).lower()
    if 'maalämpö' in a or 'maalämpöpumppu' in a:
        return 'Maalämpö'
    if 'ilmalämpö' in a or 'ilma-vesilämpöpumppu' in a:
        return 'Ilmalämpöpumppu'
    if 'kaukolämp' in a or 'kaukolämmitys' in a or 'kaukolämmön' in a:
        return 'Kaukolämpö'
    if 'öljy' in a:
        return 'Öljylämmitys'
    if any(x in a for x in ['pelletti', 'hake', 'puu', 'biomassa', 'puulämpö', 'puukattila']):
        return 'Puu / pelletti / biomassa'
    if any(x in a for x in ['sähkö', 'sahko', 'electric']):
        return 'Sähkölämmitys'
    return 'Muu / tuntematon'

def normalisoi_ilmanvaihto(arvo):
    if pd.isna(arvo):
        return 'Muu / tuntematon'
    a = str(arvo).lower()
    if 'painovoimainen' in a:
        return 'Painovoimainen'
    has_lto = 'lto' in a or 'lämmöntalteenotto' in a or 'lämmöntalteenotolla' in a
    has_tulo = 'tulo' in a
    has_poisto = 'poisto' in a or 'poistoilma' in a
    if has_tulo and has_poisto and has_lto:
        return 'Koneellinen tulo + poisto + LTO'
    if has_tulo and has_poisto and not has_lto:
        return 'Koneellinen tulo + poisto, ei LTO'
    if has_poisto and not has_tulo:
        return 'Koneellinen poisto, ei LTO'
    return 'Muu / tuntematon'

df['_lammitys_ryhma'] = df['Lahtotiedot / Lammitys / Lammitysmuoto-1 / Kuvaus-fi'].apply(normalisoi_lammitys)
df['_ilmanvaihto_ryhma'] = df['Lahtotiedot / Ilmanvaihto / Kuvaus-fi'].apply(normalisoi_ilmanvaihto)

LAMMITYS_RYHMAT = ['Kaukolämpö', 'Sähkölämmitys', 'Maalämpö', 'Ilmalämpöpumppu',
                   'Öljylämmitys', 'Puu / pelletti / biomassa', 'Muu / tuntematon']
ILMANVAIHTO_RYHMAT = ['Painovoimainen', 'Koneellinen poisto, ei LTO',
                      'Koneellinen tulo + poisto, ei LTO',
                      'Koneellinen tulo + poisto + LTO', 'Muu / tuntematon']

# Sarakealiakset
COL = {
    'id':              'Id',
    'eluokka':         'Tulokset / E-luokka',
    'eluku':           'Tulokset / E-luku',
    'kayttotarkoitus': 'Perustiedot / Kayttotarkoitus',
    'alakaytto':       'Perustiedot / Alakayttotarkoitus-fi',
    'lammitys':        'Lahtotiedot / Lammitys / Lammitysmuoto-1 / Kuvaus-fi',
    'postinumero':     'Perustiedot / Postinumero',
    'osoite':          'Perustiedot / Katuosoite-fi',
    'ilmanvaihto':     'Lahtotiedot / Ilmanvaihto / Kuvaus-fi',
    'nettoala':        'Lahtotiedot / Lammitetty-nettoala',
    'rakennustunnus':  'Perustiedot / Rakennustunnus',
    'valmistumisvuosi':'Perustiedot / Valmistumisvuosi',
    'nimi':            'Perustiedot / Nimi-fi',
    'allekirjoitus':   'Allekirjoitusaika',
    'suositukset':     'Perustiedot / Keskeiset-suositukset-fi',
    'lammitys2':       'Lahtotiedot / Lammitys / Lammitysmuoto-2 / Kuvaus-fi',
    'lammonjako':      'Lahtotiedot / Lammitys / Lammonjako / Kuvaus-fi',
}

DISPLAY_COLS = ['id','eluokka','eluku','alakaytto','lammitys','postinumero','osoite','ilmanvaihto','nettoala','valmistumisvuosi']


# ──────────────────────────────────────────────
# API-päätepisteet
# ──────────────────────────────────────────────

@app.route('/api/filters')
def get_filters():
    """Palauttaa suodatinvaihtoehdot"""
    return jsonify({
        'eluokat':        ['A','B','C','D','E','F','G'],
        'kayttotarkoitukset': sorted([
            {'koodi': koodi, 'nimi': str(nimi) if nimi else koodi}
            for koodi, nimi in df.groupby(COL['kayttotarkoitus'])[COL['alakaytto']].first().items()
            if str(koodi).strip()
        ], key=lambda x: x['nimi']),
        'lammitysmuodot':   LAMMITYS_RYHMAT,
        'ilmanvaihtotyypit': ILMANVAIHTO_RYHMAT,
        'kaupungit':        KAUPUNGIT,
    })


@app.route('/api/kaupungit')
def hae_kaupungit():
    """Hakee kaupunkeja hakusanalla (autocomplete)"""
    q = request.args.get('q', '').strip().lower()
    if not q:
        return jsonify([])
    tulokset = [k for k in KAUPUNGIT if q in k.lower()][:20]
    return jsonify(tulokset)


@app.route('/api/energiatodistukset')
def get_energiatodistukset():
    """Palauttaa suodatetut ja sivutetut tulokset"""
    filtered = df

    # Suodattimet
    eluokka_filter = request.args.getlist('eluokka')
    kaytto_filter  = request.args.getlist('kayttotarkoitus')
    lammitys_filter= request.args.getlist('lammitys')
    ilma_filter    = request.args.getlist('ilmanvaihto')
    kaupunki       = request.args.get('kaupunki', '').strip()
    ala_min        = request.args.get('ala_min', type=float)
    ala_max        = request.args.get('ala_max', type=float)
    page           = request.args.get('page', 1, type=int)
    per_page       = request.args.get('per_page', 50, type=int)
    sort_col       = request.args.get('sort', 'eluku')
    sort_dir       = request.args.get('dir', 'asc')
    omistaja_filter= request.args.get('omistaja', '').strip()
    kontaktoitu_filter = request.args.get('kontaktoitu', '').strip()  # 'kylla' | 'ei'

    if eluokka_filter:
        filtered = filtered[filtered[COL['eluokka']].isin(eluokka_filter)]
    if kaytto_filter:
        filtered = filtered[filtered[COL['kayttotarkoitus']].isin(kaytto_filter)]
    if lammitys_filter:
        filtered = filtered[filtered['_lammitys_ryhma'].isin(lammitys_filter)]
    if ilma_filter:
        filtered = filtered[filtered['_ilmanvaihto_ryhma'].isin(ilma_filter)]
    if kaupunki:
        postit = KAUPUNKI_POSTIT.get(kaupunki, [])
        if postit:
            filtered = filtered[filtered[COL['postinumero']].astype(str).str.zfill(5).isin(postit)]
        else:
            # fallback: tekstihaku postinumerolla tai kaupungin alkuosalla
            filtered = filtered[filtered[COL['postinumero']].astype(str).str.startswith(kaupunki)]
    if ala_min is not None:
        filtered = filtered[filtered[COL['nettoala']] >= ala_min]
    if ala_max is not None:
        filtered = filtered[filtered[COL['nettoala']] <= ala_max]

    # Omistaja- ja kontaktoitu-suodatus (tietokannasta)
    if omistaja_filter or kontaktoitu_filter:
        tallennetut = hae_tallennetut()
        if omistaja_filter:
            tunnukset = {k for k, v in tallennetut.items()
                         if omistaja_filter.lower() in (v.get('omistaja') or '').lower()}
            filtered = filtered[filtered[COL['rakennustunnus']].isin(tunnukset)]
        if kontaktoitu_filter == 'kylla':
            tunnukset = {k for k, v in tallennetut.items() if v.get('kontaktoitu_at')}
            filtered = filtered[filtered[COL['rakennustunnus']].isin(tunnukset)]
        elif kontaktoitu_filter == 'ei':
            tallennetut = hae_tallennetut()
            tunnukset = {k for k, v in tallennetut.items() if v.get('kontaktoitu_at')}
            filtered = filtered[~filtered[COL['rakennustunnus']].isin(tunnukset)]

    # Lajittelu
    prospektoi = request.args.get('prospektoi', '0') == '1'
    if prospektoi:
        # Prospektointijärjestys: huonoin E-luokka ensin, korkein E-luku, suurin ala, uusin vuosi
        ELUOKKA_JARJESTYS = {'G': 0, 'F': 1, 'E': 2, 'D': 3, 'C': 4, 'B': 5, 'A': 6}
        filtered = filtered.copy()
        filtered['_eluokka_rank'] = filtered[COL['eluokka']].map(ELUOKKA_JARJESTYS).fillna(99)
        try:
            filtered = filtered.sort_values(
                ['_eluokka_rank', COL['eluku'], COL['nettoala'], COL['valmistumisvuosi']],
                ascending=[True, False, False, False],
                na_position='last'
            )
        except Exception:
            pass
    else:
        sort_key = COL.get(sort_col, COL['eluku'])
        ascending = sort_dir != 'desc'
        try:
            filtered = filtered.sort_values(sort_key, ascending=ascending, na_position='last')
        except Exception:
            pass

    total = len(filtered)
    start = (page - 1) * per_page
    page_data = filtered.iloc[start:start + per_page]

    records = []
    for _, row in page_data.iterrows():
        rec = {}
        for alias, col in COL.items():
            if col in row.index:
                v = row[col]
                rec[alias] = None if pd.isna(v) else v
        records.append(rec)

    return jsonify({'total': total, 'page': page, 'per_page': per_page, 'data': records})



@app.route('/api/test-avain', methods=['GET'])
def test_avain():
    """Debug: tarkistaa onko API-avain ladattu oikein"""
    # Lue .env-tiedosto suoraan tässä kutsussa
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
    tiedosto_avain = ''
    tiedosto_ok = os.path.exists(env_path)
    if tiedosto_ok:
        with open(env_path, encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line.startswith('ANTHROPIC_API_KEY='):
                    tiedosto_avain = line.split('=', 1)[1].strip()
                    break

    key = os.environ.get('ANTHROPIC_API_KEY', '').strip()
    return jsonify({
        'ymparisto_avain': {
            'pituus': len(key),
            'alku': key[:20] + '...' if key else '',
            'muoto_ok': key.startswith('sk-ant-')
        },
        'tiedosto_avain': {
            'pituus': len(tiedosto_avain),
            'alku': tiedosto_avain[:20] + '...' if tiedosto_avain else '',
            'muoto_ok': tiedosto_avain.startswith('sk-ant-'),
            'tiedosto_loytyy': tiedosto_ok
        },
        'ohje': 'Käynnistä python app.py uudelleen jos ymparisto_avain != tiedosto_avain'
    })


@app.route('/api/omistaja-ai', methods=['POST'])
def omistaja_ai():
    """Hakee rakennuksen omistajan Claude AI:n avulla strukturoituna JSON-vastauksena."""
    data     = request.json or {}
    rakennus = data.get('rakennus', {})

    nimi           = (rakennus.get('nimi')           or '').strip()
    osoite         = (rakennus.get('osoite')         or '').strip()
    postinumero    = str(rakennus.get('postinumero') or '').strip()
    alakaytto      = (rakennus.get('alakaytto') or rakennus.get('kayttotarkoitus') or '').strip()
    valmistumisvuosi = rakennus.get('valmistumisvuosi') or ''
    nettoala       = rakennus.get('nettoala') or ''

    api_key = os.environ.get('ANTHROPIC_API_KEY', '').strip()
    if not api_key:
        return jsonify({
            'success': False,
            'error': (
                'ANTHROPIC_API_KEY puuttuu.\n\n'
                'Vaihtoehto 1 – .env-tiedosto (suositeltu):\n'
                '  1. Kopioi .env.example → .env\n'
                '  2. Lisää avain: https://console.anthropic.com/settings/keys\n'
                '  3. Käynnistä python app.py uudelleen\n\n'
                'Vaihtoehto 2 – komentorivi:\n'
                '  set ANTHROPIC_API_KEY=sk-ant-api03-... && python app.py'
            )
        })

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)

        osoite_full = f'{osoite}, {postinumero}' if postinumero else osoite
        ala_str     = f'{int(nettoala)} m²' if nettoala else 'ei tiedossa'

        prompt = f"""You are an expert system for identifying ownership and key contacts for Finnish commercial real estate.

INPUT:
- Building name: {nimi or 'unknown'}
- Address: {osoite_full}
- Building type: {alakaytto or 'unknown'}
- Built: {valmistumisvuosi or 'unknown'}
- Floor area: {ala_str}

GOAL: Determine:
1. Property identity
2. Ownership structure
3. Best matching property contact (property manager / leasing contact)

STEP 0 — KNOWN PROPERTY RECOGNITION (CRITICAL)
- Check if the property is a well-known asset
- If matched, use known owner as primary signal

STEP 1 — IDENTIFY PROPERTY
- Resolve address → property/building name

STEP 2 — FIND OPERATORS
- Identify: asset manager, leasing agent, property manager
- These are the PRIMARY sources for contact persons

STEP 3 — CONTACT DISCOVERY (HIGH PRIORITY)
Find the best available contact person related to the property.
Search sources in this order:
1. Leasing listings (toimitilat.fi, oikotie, company sites)
2. Asset manager website (e.g. JLL, Newsec, Juhola)
3. Property manager listings
4. Company portfolio pages
Select the BEST MATCH based on:
- Direct mention of the property (highest priority)
- Same building or exact address
- Same owner portfolio
Return: name, role, company, email (if available), phone (if available)
If multiple contacts exist, choose the most relevant (closest to leasing or property responsibility)

STEP 4 — OWNER IDENTIFICATION
Priority: 1. Known asset-owner match  2. Direct ownership info  3. Inference via operator

STEP 5 — CONFIDENCE
- HIGH: direct + named property + contact match
- MEDIUM: indirect but consistent signals
- LOW: weak signals

RULES:
- Prefer a real usable contact over perfect ownership certainty
- DO NOT hallucinate emails or phone numbers — leave empty if not known
- If exact person not found, return closest valid contact
- Reasoning must be in Finnish

Return ONLY valid JSON, no markdown:
{{
  "address": "{osoite_full}",
  "property_name": "",
  "recognized_asset": false,
  "owner": {{
    "name": "",
    "confidence": "HIGH | MEDIUM | LOW"
  }},
  "asset_manager": "",
  "contact": {{
    "name": "",
    "role": "",
    "company": "",
    "email": "",
    "phone": "",
    "match_quality": "DIRECT | CLOSE | PORTFOLIO"
  }},
  "evidence": [],
  "reasoning": ""
}}"""

        message = client.messages.create(
            model='claude-haiku-4-5-20251001',
            max_tokens=1500,
            messages=[{'role': 'user', 'content': prompt}]
        )

        raw = message.content[0].text.strip()
        # Poista mahdolliset markdown-koodilohkot
        raw = re.sub(r'^```(?:json)?\s*', '', raw, flags=re.MULTILINE)
        raw = re.sub(r'\s*```\s*$', '', raw, flags=re.MULTILINE)

        result = json.loads(raw)
        return jsonify({'success': True, 'result': result})

    except json.JSONDecodeError:
        return jsonify({'success': False, 'error': 'AI:n vastaus ei ollut kelvollista JSON:ia.', 'raw': raw})
    except Exception as e:
        err = str(e)
        if '401' in err or 'invalid x-api-key' in err or 'authentication_error' in err:
            return jsonify({
                'success': False,
                'error': (
                    'API-avain on virheellinen tai vanhentunut (401).\n\n'
                    'Hae uusi avain: https://console.anthropic.com/settings/keys\n'
                    'Lisää se .env-tiedostoon tai aseta ennen käynnistystä:\n'
                    '  set ANTHROPIC_API_KEY=sk-ant-api03-...'
                )
            })
        return jsonify({'success': False, 'error': err})


@app.route('/api/hyvaksy-yhtio', methods=['POST'])
def hyvaksy_yhtio():
    """Tallentaa hyväksytyn yhtiönimen tai omistajan tietokantaan"""
    data = request.json or {}
    rakennustunnus   = (data.get('rakennustunnus') or '').strip()
    osoite           = (data.get('osoite')         or '').strip()
    yhtio_nimi       = (data.get('yhtio_nimi')     or '').strip()
    omistaja         = (data.get('omistaja')        or '').strip()
    ytunnus          = (data.get('ytunnus')         or '').strip()
    tyhjenna_omistaja = bool(data.get('tyhjenna_omistaja', False))

    if not rakennustunnus or (not yhtio_nimi and not omistaja and not tyhjenna_omistaja):
        return jsonify({'success': False, 'error': 'Rakennustunnus ja nimi vaaditaan'})

    try:
        with db_lock:
            with sqlite3.connect(DB_PATH) as con:
                # Hae olemassa oleva rivi jos on
                existing = con.execute(
                    'SELECT yhtio_nimi, omistaja FROM yhtiot WHERE rakennustunnus=?',
                    (rakennustunnus,)
                ).fetchone()
                if existing:
                    new_yhtio    = yhtio_nimi or existing[0] or ''
                    # tyhjenna_omistaja=True → aseta tyhjäksi, muuten säilytä vanha jos uusi puuttuu
                    new_omistaja = '' if tyhjenna_omistaja else (omistaja or existing[1] or '')
                else:
                    new_yhtio    = yhtio_nimi or ''
                    new_omistaja = '' if tyhjenna_omistaja else omistaja
                con.execute('''INSERT OR REPLACE INTO yhtiot
                    (rakennustunnus, osoite, yhtio_nimi, omistaja, ytunnus, hyvaksytty_at)
                    VALUES (?, ?, ?, ?, ?, datetime('now','localtime'))''',
                    (rakennustunnus, osoite, new_yhtio, new_omistaja, ytunnus))
                con.commit()
        return jsonify({'success': True, 'yhtio_nimi': new_yhtio, 'omistaja': new_omistaja})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/tallennetut')
def get_tallennetut():
    """Palauttaa kaikki tallennetut yhtiönimet"""
    return jsonify(list(hae_tallennetut().values()))


@app.route('/api/merkitse-kontaktoiduksi', methods=['POST'])
def merkitse_kontaktoiduksi():
    data = request.json or {}
    rakennustunnus = (data.get('rakennustunnus') or '').strip()
    poista = bool(data.get('poista', False))
    if not rakennustunnus:
        return jsonify({'success': False, 'error': 'rakennustunnus puuttuu'})
    arvo = None if poista else __import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')
    try:
        with db_lock:
            with sqlite3.connect(DB_PATH) as con:
                existing = con.execute(
                    'SELECT rakennustunnus FROM yhtiot WHERE rakennustunnus=?', (rakennustunnus,)
                ).fetchone()
                if existing:
                    con.execute('UPDATE yhtiot SET kontaktoitu_at=? WHERE rakennustunnus=?',
                                (arvo, rakennustunnus))
                else:
                    con.execute('''INSERT INTO yhtiot (rakennustunnus, yhtio_nimi, kontaktoitu_at)
                                   VALUES (?, ?, ?)''', (rakennustunnus, '', arvo))
                con.commit()
        return jsonify({'success': True, 'kontaktoitu_at': arvo})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/tallenna-muistiinpano', methods=['POST'])
def tallenna_muistiinpano():
    data = request.json or {}
    rakennustunnus = (data.get('rakennustunnus') or '').strip()
    teksti = (data.get('teksti') or '').strip()
    if not rakennustunnus:
        return jsonify({'success': False, 'error': 'rakennustunnus puuttuu'})
    try:
        with db_lock:
            with sqlite3.connect(DB_PATH) as con:
                existing = con.execute(
                    'SELECT rakennustunnus FROM yhtiot WHERE rakennustunnus=?', (rakennustunnus,)
                ).fetchone()
                if existing:
                    con.execute('UPDATE yhtiot SET muistiinpanot=? WHERE rakennustunnus=?',
                                (teksti or None, rakennustunnus))
                else:
                    con.execute('''INSERT INTO yhtiot (rakennustunnus, yhtio_nimi, muistiinpanot)
                                   VALUES (?, ?, ?)''', (rakennustunnus, '', teksti or None))
                con.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/massahaku', methods=['POST'])
def massahaku():
    """Ajaa AI-omistajahaut useille rakennuksille kerralla (max 20)."""
    data = request.json or {}
    rakennukset = data.get('rakennukset', [])[:20]
    api_key = os.environ.get('ANTHROPIC_API_KEY', '').strip()
    if not api_key:
        return jsonify({'success': False, 'error': 'ANTHROPIC_API_KEY puuttuu'})
    if not rakennukset:
        return jsonify({'success': False, 'error': 'Ei rakennuksia'})

    import anthropic
    client = anthropic.Anthropic(api_key=api_key)
    tulokset = []

    for rakennus in rakennukset:
        nimi      = (rakennus.get('nimi') or '').strip()
        osoite    = (rakennus.get('osoite') or '').strip()
        postinumero = str(rakennus.get('postinumero') or '').strip()
        alakaytto = (rakennus.get('alakaytto') or '').strip()
        valmistumisvuosi = rakennus.get('valmistumisvuosi') or ''
        nettoala  = rakennus.get('nettoala') or ''
        rakennustunnus = (rakennus.get('rakennustunnus') or '').strip()

        # Ohita jos omistaja jo tallennettu
        tallennetut = hae_tallennetut()
        if rakennustunnus and tallennetut.get(rakennustunnus, {}).get('omistaja'):
            tulokset.append({'rakennustunnus': rakennustunnus, 'skip': True,
                             'omistaja': tallennetut[rakennustunnus]['omistaja']})
            continue

        osoite_full = f'{osoite}, {postinumero}' if postinumero else osoite
        ala_str = f'{int(nettoala)} m²' if nettoala else 'ei tiedossa'

        prompt = f"""Selvitä suomalaisen kiinteistön omistaja. Palauta VAIN JSON.
Rakennus: {nimi or 'ei tiedossa'}, {osoite_full}, {alakaytto or ''}, {valmistumisvuosi or ''}, {ala_str}
MUISTA: omistaja ≠ vuokralainen. Anna LOW confidence jos epävarma.
{{"owner":{{"name":"","confidence":"HIGH|MEDIUM|LOW"}},"reasoning":""}}"""

        try:
            msg = client.messages.create(
                model='claude-haiku-4-5-20251001', max_tokens=256,
                messages=[{'role': 'user', 'content': prompt}]
            )
            raw = msg.content[0].text.strip()
            raw = re.sub(r'^```(?:json)?\s*', '', raw, flags=re.MULTILINE)
            raw = re.sub(r'\s*```\s*$', '', raw, flags=re.MULTILINE)
            result = json.loads(raw)
            owner_name = result.get('owner', {}).get('name', '')
            confidence = result.get('owner', {}).get('confidence', 'LOW')

            # Tallenna automaattisesti jos HIGH tai MEDIUM
            if owner_name and confidence in ('HIGH', 'MEDIUM') and rakennustunnus:
                with db_lock:
                    with sqlite3.connect(DB_PATH) as con:
                        existing = con.execute(
                            'SELECT yhtio_nimi FROM yhtiot WHERE rakennustunnus=?',
                            (rakennustunnus,)).fetchone()
                        if existing:
                            con.execute('UPDATE yhtiot SET omistaja=? WHERE rakennustunnus=?',
                                        (owner_name, rakennustunnus))
                        else:
                            con.execute('''INSERT INTO yhtiot (rakennustunnus, osoite, yhtio_nimi, omistaja)
                                           VALUES (?,?,?,?)''',
                                        (rakennustunnus, osoite, nimi, owner_name))
                        con.commit()

            tulokset.append({
                'rakennustunnus': rakennustunnus, 'nimi': nimi, 'osoite': osoite,
                'omistaja': owner_name, 'confidence': confidence,
                'reasoning': result.get('reasoning', ''), 'tallennettu': confidence in ('HIGH', 'MEDIUM')
            })
        except Exception as e:
            tulokset.append({'rakennustunnus': rakennustunnus, 'nimi': nimi,
                             'osoite': osoite, 'virhe': str(e)})

    return jsonify({'success': True, 'tulokset': tulokset})


@app.route('/api/vie-excel')
def vie_excel():
    """Vie tallennetut omistajat + rakennustiedot Excel-tiedostona."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import Response

        tallennetut = hae_tallennetut()
        if not tallennetut:
            return jsonify({'error': 'Ei tallennettuja tietoja'}), 404

        tunnukset = set(tallennetut.keys())
        rak_col = COL.get('rakennustunnus', '')
        if rak_col:
            mask = df[rak_col].isin(tunnukset)
            subset = df[mask].copy()
        else:
            subset = pd.DataFrame()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Omistajat'

        headers = ['Rakennustunnus', 'Nimi', 'Osoite', 'Postinumero',
                   'E-luokka', 'E-luku', 'Pinta-ala (m²)', 'Valmistumisvuosi',
                   'Käyttötarkoitus', 'Omistaja', 'Yhtiönimi', 'Kontaktoitu', 'Muistiinpanot']
        header_fill = PatternFill('solid', fgColor='1e3a5f')
        header_font = Font(bold=True, color='FFFFFF')

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        col_keys = ['rakennustunnus', 'nimi', 'osoite', 'postinumero',
                    'eluokka', 'eluku', 'nettoala', 'valmistumisvuosi', 'alakaytto']

        for ri, (_, row) in enumerate(subset.iterrows(), 2):
            tunnus = str(row.get(COL.get('rakennustunnus', ''), '') or '')
            t = tallennetut.get(tunnus, {})
            for ci, key in enumerate(col_keys, 1):
                v = row.get(COL.get(key, ''), '')
                ws.cell(row=ri, column=ci, value=None if pd.isna(v) else v)
            ws.cell(row=ri, column=10, value=t.get('omistaja', ''))
            ws.cell(row=ri, column=11, value=t.get('yhtio_nimi', ''))
            ws.cell(row=ri, column=12, value=t.get('kontaktoitu_at', ''))
            ws.cell(row=ri, column=13, value=t.get('muistiinpanot', ''))

        # Sarakeleveydet
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=omistajat.xlsx'}
        )
    except ImportError:
        return jsonify({'error': 'openpyxl puuttuu — lisää requirements.txt:ään'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ──────────────────────────────────────────────
# Frontend (HTML)
# ──────────────────────────────────────────────

@app.route('/')
def index():
    return render_template_string(HTML)


HTML = r"""<!DOCTYPE html>
<html lang="fi">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Energiatodistusrekisteri</title>
<script src="https://cdn.tailwindcss.com"></script>
<style>
  body { font-family: 'Inter', system-ui, sans-serif; background: #f8fafc; }
  .eluokka-A { background:#16a34a; color:#fff; }
  .eluokka-B { background:#65a30d; color:#fff; }
  .eluokka-C { background:#ca8a04; color:#fff; }
  .eluokka-D { background:#d97706; color:#fff; }
  .eluokka-E { background:#ea580c; color:#fff; }
  .eluokka-F { background:#dc2626; color:#fff; }
  .eluokka-G { background:#7f1d1d; color:#fff; }
  .filter-chip { cursor:pointer; user-select:none; transition:all .15s; }
  .filter-chip.active { ring: 2px; }
  .scrollbar-thin::-webkit-scrollbar { width:4px; }
  .scrollbar-thin::-webkit-scrollbar-thumb { background:#cbd5e1; border-radius:2px; }
  .card-selected { outline: 2px solid #3b82f6; outline-offset: -2px; }
  .bg-eluokka-A{background:#16a34a} .bg-eluokka-B{background:#65a30d} .bg-eluokka-C{background:#ca8a04}
  .bg-eluokka-D{background:#d97706} .bg-eluokka-E{background:#ea580c} .bg-eluokka-F{background:#dc2626}
  .bg-eluokka-G{background:#7f1d1d}
  .border-eluokka-A{border-left-color:#16a34a} .border-eluokka-B{border-left-color:#65a30d}
  .border-eluokka-C{border-left-color:#ca8a04} .border-eluokka-D{border-left-color:#d97706}
  .border-eluokka-E{border-left-color:#ea580c} .border-eluokka-F{border-left-color:#dc2626}
  .border-eluokka-G{border-left-color:#7f1d1d}
  #modal { display:none; }
  #modal.open { display:flex; }
  .loader { border:3px solid #e2e8f0; border-top-color:#3b82f6; border-radius:50%; width:28px; height:28px; animation:spin .7s linear infinite; }
  @keyframes spin { to { transform:rotate(360deg); } }
</style>
</head>
<body class="min-h-screen">

<!-- ── HEADER ── -->
<header class="bg-white border-b border-slate-200 px-6 py-4 flex items-center gap-4 sticky top-0 z-20 shadow-sm">
  <div class="flex items-center gap-2">
    <svg class="w-7 h-7 text-blue-600" fill="none" stroke="currentColor" viewBox="0 0 24 24">
      <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M9 12l2 2 4-4m5.618-4.016A11.955 11.955 0 0112 2.944a11.955 11.955 0 01-8.618 3.04A12.02 12.02 0 003 9c0 5.591 3.824 10.29 9 11.622 5.176-1.332 9-6.03 9-11.622 0-1.042-.133-2.052-.382-3.016z"/>
    </svg>
    <span class="font-bold text-xl text-slate-800">Energiatodistusrekisteri</span>
  </div>
  <div class="ml-auto text-sm text-slate-500" id="total-label">Ladataan...</div>
</header>

<!-- ── LAYOUT ── -->
<div class="flex h-[calc(100vh-65px)]">

  <!-- SIDEBAR -->
  <aside class="w-72 bg-white border-r border-slate-200 flex flex-col overflow-hidden flex-shrink-0">
    <div class="px-4 py-3 border-b border-slate-100 flex items-center justify-between">
      <span class="font-semibold text-slate-700 text-sm">Suodattimet</span>
      <button onclick="resetFilters()" class="text-xs text-blue-600 hover:underline">Tyhjennä</button>
    </div>
    <div class="flex-1 overflow-y-auto scrollbar-thin p-4 space-y-5">

      <!-- E-luokka -->
      <div>
        <label class="block text-xs font-semibold text-slate-500 uppercase tracking-wider mb-2">E-luokka</label>
        <div class="flex flex-wrap gap-1.5" id="eluokka-chips"></div>
      </div>

      <!-- Rakennustyyppi -->
      <div>
        <label class="block text-xs font-semibold text-slate-500 uppercase tracking-wider mb-2">Rakennustyyppi</label>
        <select id="kayttotarkoitus-select" multiple class="w-full text-sm border border-slate-200 rounded-lg p-2 h-36 focus:ring-2 focus:ring-blue-500 focus:outline-none"></select>
        <div class="text-xs text-slate-400 mt-1">Ctrl+klikkaa useita</div>
      </div>

      <!-- Lämmitysmuoto -->
      <div>
        <label class="block text-xs font-semibold text-slate-500 uppercase tracking-wider mb-2">Lämmitysmuoto</label>
        <select id="lammitys-select" multiple class="w-full text-sm border border-slate-200 rounded-lg p-2 h-32 focus:ring-2 focus:ring-blue-500 focus:outline-none"></select>
        <div class="text-xs text-slate-400 mt-1">Ctrl+klikkaa useita</div>
      </div>

      <!-- Ilmanvaihtotapa -->
      <div>
        <label class="block text-xs font-semibold text-slate-500 uppercase tracking-wider mb-2">Ilmanvaihtotapa</label>
        <select id="ilmanvaihto-select" multiple class="w-full text-sm border border-slate-200 rounded-lg p-2 h-28 focus:ring-2 focus:ring-blue-500 focus:outline-none"></select>
        <div class="text-xs text-slate-400 mt-1">Ctrl+klikkaa useita</div>
      </div>

      <!-- Kaupunki -->
      <div class="relative">
        <label class="block text-xs font-semibold text-slate-500 uppercase tracking-wider mb-2">Kaupunki</label>
        <input id="kaupunki-input" type="text" placeholder="Hae kaupunkia..." autocomplete="off"
          class="w-full text-sm border border-slate-200 rounded-lg px-3 py-2 focus:ring-2 focus:ring-blue-500 focus:outline-none">
        <div id="kaupunki-dropdown" class="hidden absolute z-30 w-full bg-white border border-slate-200 rounded-lg shadow-lg mt-1 max-h-48 overflow-y-auto scrollbar-thin"></div>
      </div>

      <!-- Pinta-ala -->
      <div>
        <label class="block text-xs font-semibold text-slate-500 uppercase tracking-wider mb-2">Pinta-ala (m²)</label>
        <div class="flex gap-2">
          <input id="ala-min" type="number" placeholder="Min" min="0"
            class="w-full text-sm border border-slate-200 rounded-lg px-3 py-2 focus:ring-2 focus:ring-blue-500 focus:outline-none">
          <input id="ala-max" type="number" placeholder="Max"
            class="w-full text-sm border border-slate-200 rounded-lg px-3 py-2 focus:ring-2 focus:ring-blue-500 focus:outline-none">
        </div>
      </div>

      <!-- Omistaja-haku -->
      <div>
        <label class="block text-xs font-semibold text-slate-500 uppercase tracking-wider mb-2">Omistaja</label>
        <input id="omistaja-filter" type="text" placeholder="Hae omistajalla..."
          class="w-full text-sm border border-slate-200 rounded-lg px-3 py-2 focus:ring-2 focus:ring-violet-400 focus:outline-none">
      </div>

      <!-- Kontaktoitu -->
      <div>
        <label class="block text-xs font-semibold text-slate-500 uppercase tracking-wider mb-2">Kontaktoitu</label>
        <select id="kontaktoitu-filter"
          class="w-full text-sm border border-slate-200 rounded-lg px-3 py-2 focus:ring-2 focus:ring-blue-500 focus:outline-none">
          <option value="">Kaikki</option>
          <option value="ei">Ei kontaktoitu</option>
          <option value="kylla">Kontaktoitu</option>
        </select>
      </div>

    </div>
    <div class="p-4 border-t border-slate-100">
      <button onclick="applyFilters()" class="w-full bg-blue-600 text-white font-medium text-sm py-2.5 rounded-lg hover:bg-blue-700 transition-colors">
        Hae tulokset
      </button>
    </div>
  </aside>

  <!-- MAIN -->
  <main class="flex-1 flex flex-col overflow-hidden">

    <!-- Table toolbar -->
    <div class="bg-white border-b border-slate-200 px-4 py-2 flex items-center gap-3">
      <span class="text-sm text-slate-600" id="result-count"></span>
      <button id="prospektoi-btn" onclick="toggleProspektoi()"
        class="flex items-center gap-1.5 text-sm font-medium px-3 py-1.5 rounded-lg border border-slate-200 hover:border-amber-400 hover:bg-amber-50 hover:text-amber-700 transition-colors">
        <svg class="w-4 h-4" fill="none" stroke="currentColor" viewBox="0 0 24 24">
          <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2"
            d="M13 7h8m0 0v8m0-8l-8 8-4-4-6 6"/>
        </svg>
        Prospektoi
      </button>
      <button onclick="kaynnistaMassahaku()"
        class="flex items-center gap-1.5 text-sm font-medium px-3 py-1.5 rounded-lg border border-slate-200 hover:border-violet-400 hover:bg-violet-50 hover:text-violet-700 transition-colors">
        <svg class="w-4 h-4" fill="none" stroke="currentColor" viewBox="0 0 24 24">
          <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M19 11H5m14 0a2 2 0 012 2v6a2 2 0 01-2 2H5a2 2 0 01-2-2v-6a2 2 0 012-2m14 0V9a2 2 0 00-2-2M5 11V9a2 2 0 012-2m0 0V5a2 2 0 012-2h6a2 2 0 012 2v2M7 7h10"/>
        </svg>
        Massahaku
      </button>
      <a href="/api/vie-excel" download="omistajat.xlsx"
        class="flex items-center gap-1.5 text-sm font-medium px-3 py-1.5 rounded-lg border border-slate-200 hover:border-emerald-400 hover:bg-emerald-50 hover:text-emerald-700 transition-colors">
        <svg class="w-4 h-4" fill="none" stroke="currentColor" viewBox="0 0 24 24">
          <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M12 10v6m0 0l-3-3m3 3l3-3m2 8H7a2 2 0 01-2-2V5a2 2 0 012-2h5.586a1 1 0 01.707.293l5.414 5.414a1 1 0 01.293.707V19a2 2 0 01-2 2z"/>
        </svg>
        Vie Excel
      </a>
      <div id="massahaku-status" class="hidden text-xs text-violet-600 bg-violet-50 px-2 py-1 rounded-lg"></div>
      <div class="ml-auto flex items-center gap-3">
        <select id="sort-select" onchange="handleSortSelect()" class="text-sm border border-slate-200 rounded-lg px-2 py-1.5 text-slate-600">
          <option value="eluku|desc">E-luku ↓ ensin</option>
          <option value="eluku|asc">E-luku ↑ ensin</option>
          <option value="nettoala|desc">Suurin ala ensin</option>
          <option value="nettoala|asc">Pienin ala ensin</option>
          <option value="valmistumisvuosi|desc">Uusin ensin</option>
          <option value="valmistumisvuosi|asc">Vanhin ensin</option>
          <option value="eluokka|asc">E-luokka A→G</option>
          <option value="eluokka|desc">E-luokka G→A</option>
          <option value="nimi|asc">Nimi A→Ö</option>
          <option value="nimi|desc">Nimi Ö→A</option>
        </select>
        <span class="text-slate-300">|</span>
        <span class="text-sm text-slate-500">Näytä</span>
        <select id="per-page" onchange="applyFilters()" class="text-sm border border-slate-200 rounded-lg px-2 py-1.5">
          <option value="25">25</option>
          <option value="50" selected>50</option>
          <option value="100">100</option>
        </select>
      </div>
    </div>

    <!-- Korttilistaus -->
    <div class="flex-1 overflow-auto scrollbar-thin">
      <div id="table-body" class="p-4 space-y-2"></div>
      <div id="loading" class="hidden flex items-center justify-center py-16">
        <div class="loader"></div>
      </div>
      <div id="empty-state" class="hidden text-center py-16 text-slate-400">
        <svg class="w-10 h-10 mx-auto mb-3 opacity-40" fill="none" stroke="currentColor" viewBox="0 0 24 24">
          <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M9.172 16.172a4 4 0 015.656 0M9 10h.01M15 10h.01M21 12a9 9 0 11-18 0 9 9 0 0118 0z"/>
        </svg>
        <p>Ei tuloksia valituilla suodattimilla</p>
      </div>
    </div>

    <!-- Pagination -->
    <div class="bg-white border-t border-slate-200 px-4 py-3 flex items-center justify-between">
      <button id="prev-btn" onclick="changePage(-1)" class="text-sm px-4 py-1.5 rounded-lg border border-slate-200 hover:bg-slate-50 disabled:opacity-40 disabled:cursor-not-allowed transition-colors">
        ← Edellinen
      </button>
      <span class="text-sm text-slate-600" id="page-info"></span>
      <button id="next-btn" onclick="changePage(1)" class="text-sm px-4 py-1.5 rounded-lg border border-slate-200 hover:bg-slate-50 disabled:opacity-40 disabled:cursor-not-allowed transition-colors">
        Seuraava →
      </button>
    </div>
  </main>
</div>

<!-- ── DETAIL MODAL ── -->
<div id="modal" class="fixed inset-0 z-50 items-center justify-center bg-black/40 backdrop-blur-sm p-4">
  <div class="bg-white rounded-2xl shadow-2xl w-full max-w-2xl max-h-[90vh] flex flex-col">

    <!-- Modal header -->
    <div class="px-6 py-4 border-b border-slate-100 flex items-start justify-between">
      <div>
        <h2 class="font-bold text-lg text-slate-800" id="modal-osoite">—</h2>
        <p class="text-sm text-slate-500" id="modal-tyyppi">—</p>
      </div>
      <button onclick="closeModal()" class="text-slate-400 hover:text-slate-600 ml-4 flex-shrink-0">
        <svg class="w-6 h-6" fill="none" stroke="currentColor" viewBox="0 0 24 24">
          <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M6 18L18 6M6 6l12 12"/>
        </svg>
      </button>
    </div>

    <!-- Modal content -->
    <div class="flex-1 overflow-y-auto scrollbar-thin p-6 space-y-5">

      <!-- Perustiedot grid -->
      <div class="grid grid-cols-2 gap-3" id="modal-grid"></div>

      <!-- Nimi datasta -->
      <div id="nimi-ehdotus-block"></div>

      <!-- Omistajahaku -->
      <div class="border border-violet-100 rounded-xl p-4 bg-violet-50/40">

        <!-- Tallennettu omistaja (näkyy jos on) -->
        <div id="omistaja-nykyinen" class="hidden mb-3 bg-emerald-50 border border-emerald-200 rounded-lg px-3 py-2.5 flex items-center justify-between gap-2">
          <div class="flex items-center gap-2 min-w-0">
            <svg class="w-4 h-4 text-emerald-600 flex-shrink-0" fill="none" stroke="currentColor" viewBox="0 0 24 24">
              <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M5 13l4 4L19 7"/>
            </svg>
            <span class="text-sm font-semibold text-emerald-800 truncate" id="omistaja-nykyinen-nimi"></span>
          </div>
          <div class="flex gap-1.5 flex-shrink-0">
            <button onclick="korvaaOmistaja()" class="text-xs bg-white border border-slate-300 text-slate-600 px-2.5 py-1 rounded-lg hover:bg-slate-50 transition-colors">Korvaa</button>
            <button onclick="tyhjennaOmistaja()" class="text-xs bg-white border border-red-200 text-red-500 px-2.5 py-1 rounded-lg hover:bg-red-50 transition-colors">Tyhjennä</button>
          </div>
        </div>

        <div class="flex items-center justify-between gap-2 mb-3">
          <div class="flex items-center gap-2">
            <svg class="w-5 h-5 text-violet-500" fill="none" stroke="currentColor" viewBox="0 0 24 24">
              <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M16 7a4 4 0 11-8 0 4 4 0 018 0zM12 14a7 7 0 00-7 7h14a7 7 0 00-7-7z"/>
            </svg>
            <span class="font-semibold text-slate-700">Etsi omistaja</span>
          </div>
          <button id="ai-haku-btn" onclick="haeOmistajaAI()"
            class="flex items-center gap-1.5 text-xs font-medium px-3 py-1.5 rounded-lg bg-gradient-to-r from-violet-600 to-indigo-600 text-white hover:from-violet-700 hover:to-indigo-700 transition-all shadow-sm whitespace-nowrap">
            <svg class="w-3.5 h-3.5" fill="none" stroke="currentColor" viewBox="0 0 24 24">
              <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M9.663 17h4.673M12 3v1m6.364 1.636l-.707.707M21 12h-1M4 12H3m3.343-5.657l-.707-.707m2.828 9.9a5 5 0 117.072 0l-.347.346A3.51 3.51 0 0114.5 20.5H10a3.51 3.51 0 01-2.121-.754l-.347-.346z"/>
            </svg>
            AI-haku
          </button>
        </div>
        <div id="ai-loading" class="hidden flex items-center gap-3 py-2 mb-3">
          <div class="loader"></div>
          <span class="text-sm text-slate-500">Claude analysoi omistajuutta...</span>
        </div>
        <div id="ai-result" class="hidden mb-3"></div>

        <!-- Manuaalinen syöttö -->
        <div class="flex gap-2 items-center mt-1">
          <input id="omistaja-manuaali" type="text" placeholder="Kirjoita omistaja käsin..."
            class="flex-1 text-sm border border-slate-200 rounded-lg px-3 py-2 focus:ring-2 focus:ring-violet-400 focus:outline-none bg-white">
          <button onclick="tallennaManuaaliOmistaja()"
            class="text-xs bg-emerald-600 text-white px-3 py-2 rounded-lg hover:bg-emerald-700 whitespace-nowrap">
            Tallenna
          </button>
        </div>

        <!-- Tallennusvahvistus -->
        <div id="prh-result" class="hidden mt-3">
          <div id="prh-lista" class="space-y-1.5"></div>
        </div>

        <div id="omistaja-error" class="hidden text-xs text-amber-700 bg-amber-50 rounded-lg p-2 mt-2"></div>
      </div>

      <!-- Kontaktoitu -->
      <div class="border border-slate-100 rounded-xl p-4 bg-slate-50/40">
        <div class="flex items-center justify-between gap-3">
          <div>
            <div class="font-semibold text-slate-700 text-sm">Kontaktoitu</div>
            <div id="kontaktoitu-pvm" class="text-xs text-slate-400 mt-0.5"></div>
          </div>
          <button id="kontaktoitu-btn" onclick="toggleKontaktoitu()"
            class="text-sm font-medium px-4 py-2 rounded-lg border transition-colors">
          </button>
        </div>
      </div>

      <!-- Muistiinpanot -->
      <div class="border border-slate-100 rounded-xl p-4 bg-slate-50/40">
        <div class="font-semibold text-slate-700 text-sm mb-2">Muistiinpanot</div>
        <textarea id="muistiinpanot-input" rows="3" placeholder="Lisää muistiinpanoja..."
          class="w-full text-sm border border-slate-200 rounded-lg px-3 py-2 focus:ring-2 focus:ring-slate-400 focus:outline-none bg-white resize-none"></textarea>
        <button onclick="tallennaMuistiinpano()"
          class="mt-2 text-xs bg-slate-700 text-white px-3 py-1.5 rounded-lg hover:bg-slate-800 transition-colors">
          Tallenna muistiinpano
        </button>
      </div>

    </div>
  </div>
</div>

<script>
// ── STATE ──
let state = {
  page: 1,
  total: 0,
  perPage: 50,
  sort: 'eluku',
  dir: 'asc',
  eluokka: [],
  selectedRow: null,
  tallennetut: {},      // rakennustunnus → {yhtio_nimi, ytunnus, ...}
  valittuKaupunki: '',
  prospektoi: false,
};

// ── INIT ──
async function init() {
  // Lataa tallennetut yhtiönimet
  const tallennetutRes = await fetch('/api/tallennetut');
  const tallennetutLista = await tallennetutRes.json();
  tallennetutLista.forEach(t => { state.tallennetut[t.rakennustunnus] = t; });

  const res = await fetch('/api/filters');
  const filters = await res.json();

  // E-luokka chips
  const chips = document.getElementById('eluokka-chips');
  filters.eluokat.forEach(klass => {
    const chip = document.createElement('button');
    chip.className = `filter-chip eluokka-${klass} px-3 py-1 rounded-full text-sm font-bold opacity-40 hover:opacity-80`;
    chip.textContent = klass;
    chip.dataset.klass = klass;
    chip.addEventListener('click', () => toggleEluokka(klass, chip));
    chips.appendChild(chip);
  });

  // Rakennustyyppi
  const ktSel = document.getElementById('kayttotarkoitus-select');
  filters.kayttotarkoitukset.forEach(({koodi, nimi}) => {
    const opt = new Option(nimi || koodi, koodi);
    ktSel.add(opt);
  });

  // Lämmitys
  const lSel = document.getElementById('lammitys-select');
  filters.lammitysmuodot.forEach(m => lSel.add(new Option(m, m)));

  // Ilmanvaihto
  const iSel = document.getElementById('ilmanvaihto-select');
  filters.ilmanvaihtotyypit.forEach(t => iSel.add(new Option(t, t)));

  // Enter triggers search
  ['ala-min','ala-max'].forEach(id => {
    document.getElementById(id).addEventListener('keydown', e => {
      if (e.key === 'Enter') applyFilters();
    });
  });

  // Kaupunki autocomplete
  setupKaupunkiAutocomplete();

  // Sort select
  updateSortSelect();

  applyFilters();
}

function toggleEluokka(klass, chip) {
  const idx = state.eluokka.indexOf(klass);
  if (idx >= 0) {
    state.eluokka.splice(idx, 1);
    chip.classList.add('opacity-40');
    chip.classList.remove('opacity-100', 'ring-2', 'ring-white', 'ring-offset-1');
  } else {
    state.eluokka.push(klass);
    chip.classList.remove('opacity-40');
    chip.classList.add('opacity-100', 'ring-2', 'ring-white', 'ring-offset-1');
  }
}

function applyFilters() {
  state.page = 1;
  fetchData();
}

function toggleProspektoi() {
  state.prospektoi = !state.prospektoi;
  state.page = 1;
  const btn = document.getElementById('prospektoi-btn');
  if (state.prospektoi) {
    btn.classList.add('bg-amber-500', 'text-white', 'border-amber-500');
    btn.classList.remove('border-slate-200', 'hover:border-amber-400', 'hover:bg-amber-50', 'hover:text-amber-700');
  } else {
    btn.classList.remove('bg-amber-500', 'text-white', 'border-amber-500');
    btn.classList.add('border-slate-200', 'hover:border-amber-400', 'hover:bg-amber-50', 'hover:text-amber-700');
  }
  fetchData();
}

function changePage(delta) {
  const maxPage = Math.ceil(state.total / state.perPage);
  state.page = Math.max(1, Math.min(state.page + delta, maxPage));
  fetchData();
}

function resetFilters() {
  state.eluokka = [];
  document.querySelectorAll('#eluokka-chips .filter-chip').forEach(c => {
    c.classList.add('opacity-40');
    c.classList.remove('opacity-100','ring-2','ring-white','ring-offset-1');
  });
  ['kayttotarkoitus-select','lammitys-select','ilmanvaihto-select'].forEach(id => {
    Array.from(document.getElementById(id).options).forEach(o => o.selected = false);
  });
  document.getElementById('kaupunki-input').value = '';
  state.valittuKaupunki = '';
  document.getElementById('ala-min').value = '';
  document.getElementById('ala-max').value = '';
  applyFilters();
}

function setupKaupunkiAutocomplete() {
  const input = document.getElementById('kaupunki-input');
  const dropdown = document.getElementById('kaupunki-dropdown');
  state.valittuKaupunki = '';
  let debounceTimer;

  input.addEventListener('input', () => {
    state.valittuKaupunki = '';
    clearTimeout(debounceTimer);
    const q = input.value.trim();
    if (!q) { dropdown.classList.add('hidden'); return; }
    debounceTimer = setTimeout(async () => {
      const res = await fetch('/api/kaupungit?q=' + encodeURIComponent(q));
      const tulokset = await res.json();
      dropdown.innerHTML = '';
      if (!tulokset.length) { dropdown.classList.add('hidden'); return; }
      tulokset.forEach(kaupunki => {
        const item = document.createElement('div');
        item.className = 'px-3 py-2 text-sm text-slate-700 hover:bg-blue-50 cursor-pointer';
        item.textContent = kaupunki;
        item.addEventListener('mousedown', () => {
          input.value = kaupunki;
          state.valittuKaupunki = kaupunki;
          dropdown.classList.add('hidden');
          applyFilters();
        });
        dropdown.appendChild(item);
      });
      dropdown.classList.remove('hidden');
    }, 200);
  });

  input.addEventListener('keydown', e => {
    if (e.key === 'Escape') dropdown.classList.add('hidden');
    if (e.key === 'Enter') { dropdown.classList.add('hidden'); applyFilters(); }
  });

  document.addEventListener('click', e => {
    if (!input.contains(e.target) && !dropdown.contains(e.target)) {
      dropdown.classList.add('hidden');
    }
  });
}

function handleSortSelect() {
  const val = document.getElementById('sort-select').value;
  const [col, dir] = val.split('|');
  state.sort = col;
  state.dir  = dir;
  state.page = 1;
  fetchData();
}

function updateSortSelect() {
  const sel = document.getElementById('sort-select');
  if (sel) sel.value = `${state.sort}|${state.dir}`;
}

async function fetchData() {
  const perPage = parseInt(document.getElementById('per-page').value);
  state.perPage = perPage;

  const params = new URLSearchParams({
    page: state.page,
    per_page: perPage,
    sort: state.sort,
    dir: state.dir,
    prospektoi: state.prospektoi ? '1' : '0',
  });

  state.eluokka.forEach(e => params.append('eluokka', e));

  const ktSel = document.getElementById('kayttotarkoitus-select');
  Array.from(ktSel.selectedOptions).forEach(o => params.append('kayttotarkoitus', o.value));

  const lSel = document.getElementById('lammitys-select');
  Array.from(lSel.selectedOptions).forEach(o => params.append('lammitys', o.value));

  const iSel = document.getElementById('ilmanvaihto-select');
  Array.from(iSel.selectedOptions).forEach(o => params.append('ilmanvaihto', o.value));

  const kaupunki = state.valittuKaupunki || document.getElementById('kaupunki-input').value.trim();
  if (kaupunki) params.set('kaupunki', kaupunki);

  const alaMin = document.getElementById('ala-min').value;
  const alaMax = document.getElementById('ala-max').value;
  if (alaMin) params.set('ala_min', alaMin);
  if (alaMax) params.set('ala_max', alaMax);

  const omistajaf = document.getElementById('omistaja-filter').value.trim();
  if (omistajaf) params.set('omistaja', omistajaf);
  const kontaktoiduf = document.getElementById('kontaktoitu-filter').value;
  if (kontaktoiduf) params.set('kontaktoitu', kontaktoiduf);

  document.getElementById('loading').classList.remove('hidden');
  document.getElementById('table-body').innerHTML = '';
  document.getElementById('empty-state').classList.add('hidden');

  const res = await fetch('/api/energiatodistukset?' + params.toString());
  const data = await res.json();

  document.getElementById('loading').classList.add('hidden');
  state.total = data.total;

  renderTable(data.data);
  updatePagination(data.total, data.page, perPage);
}

function elukuColor(v) {
  if (v == null) return '#94a3b8';
  if (v >= 300) return '#b91c1c';
  if (v >= 200) return '#c2410c';
  if (v >= 150) return '#b45309';
  return '#475569';
}

function renderTable(rows) {
  const container = document.getElementById('table-body');
  container.innerHTML = '';
  document.getElementById('empty-state').classList.add('hidden');

  if (!rows.length) {
    document.getElementById('empty-state').classList.remove('hidden');
    return;
  }

  rows.forEach(row => {
    const tallennettu = state.tallennetut[row.rakennustunnus];
    const nimi        = row.nimi || '';
    const omistaja    = tallennettu?.omistaja || '';
    const el          = row.eluokka || '?';

    const kontaktoitu = tallennettu?.kontaktoitu_at || '';
    const chips = [];
    if (row.nettoala)         chips.push(`<span class="text-xs bg-slate-100 text-slate-600 px-2 py-0.5 rounded-full">${Math.round(row.nettoala).toLocaleString('fi')} m²</span>`);
    if (row.valmistumisvuosi) chips.push(`<span class="text-xs bg-slate-100 text-slate-600 px-2 py-0.5 rounded-full">${row.valmistumisvuosi}</span>`);
    if (row.lammitys)         chips.push(`<span class="text-xs bg-slate-100 text-slate-600 px-2 py-0.5 rounded-full">${row.lammitys}</span>`);
    if (omistaja)             chips.push(`<span class="omistaja-chip text-xs bg-violet-100 text-violet-700 px-2 py-0.5 rounded-full font-medium">${omistaja}</span>`);
    if (kontaktoitu)          chips.push(`<span class="kontaktoitu-chip text-xs bg-emerald-100 text-emerald-700 px-2 py-0.5 rounded-full font-medium">✓ Kontaktoitu</span>`);

    const card = document.createElement('div');
    card.className = `bg-white rounded-xl shadow-sm border-l-4 border-eluokka-${el} flex items-stretch cursor-pointer hover:shadow-md transition-all group`;
    card.dataset.tunnus = row.rakennustunnus || '';
    card.innerHTML = `
      <div class="bg-eluokka-${el} w-12 flex-shrink-0 rounded-l-xl flex items-center justify-center">
        <span class="text-white font-black text-base">${el}</span>
      </div>
      <div class="flex-1 px-4 py-3 min-w-0">
        <div class="flex items-start justify-between gap-3">
          <div class="min-w-0">
            <div class="font-semibold text-slate-800 text-sm truncate">${nimi || row.osoite || '—'}</div>
            <div class="text-xs text-slate-500 mt-0.5 truncate">${row.osoite || ''}${row.postinumero ? ' · ' + row.postinumero : ''}${row.alakaytto ? ' · ' + row.alakaytto : ''}</div>
          </div>
          <div class="text-right flex-shrink-0 ml-2">
            <div class="text-sm font-mono font-bold" style="color:${elukuColor(row.eluku)}">${row.eluku != null ? row.eluku : '—'}</div>
            <div class="text-xs text-slate-400">kWh/(m²a)</div>
          </div>
        </div>
        ${chips.length ? `<div class="flex gap-1.5 mt-2 flex-wrap">${chips.join('')}</div>` : ''}
      </div>
      <div class="flex items-center pr-3 text-slate-200 group-hover:text-slate-400 transition-colors flex-shrink-0">
        <svg class="w-4 h-4" fill="none" stroke="currentColor" viewBox="0 0 24 24">
          <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M9 5l7 7-7 7"/>
        </svg>
      </div>
    `;
    card._rowData = row;
    card.addEventListener('click', () => openModal(row, card));
    container.appendChild(card);
  });
}

function updatePagination(total, page, perPage) {
  const maxPage = Math.max(1, Math.ceil(total / perPage));
  document.getElementById('total-label').textContent = `${total.toLocaleString('fi')} todistusta`;
  document.getElementById('result-count').textContent = `${total.toLocaleString('fi')} tulosta`;
  document.getElementById('page-info').textContent = `Sivu ${page} / ${maxPage}`;
  document.getElementById('prev-btn').disabled = page <= 1;
  document.getElementById('next-btn').disabled = page >= maxPage;
}

// ── MODAL ──
let currentRow = null;

function openModal(row, tr) {
  currentRow = row;

  if (state.selectedRow) state.selectedRow.classList.remove('card-selected');
  tr.classList.add('card-selected');
  state.selectedRow = tr;

  document.getElementById('modal-osoite').textContent =
    (row.osoite || 'Osoite ei tiedossa') + (row.postinumero ? ', ' + row.postinumero : '');
  document.getElementById('modal-tyyppi').textContent =
    row.alakaytto || row.kayttotarkoitus || '';

  const grid = document.getElementById('modal-grid');
  const fields = [
    ['E-luokka', `<span class="eluokka-${row.eluokka} text-sm font-bold px-2 py-0.5 rounded">${row.eluokka || '—'}</span>`],
    ['E-luku', row.eluku != null ? `<span class="font-mono">${row.eluku} kWh/(m²a)</span>` : '—'],
    ['Pinta-ala', row.nettoala != null ? `${Math.round(row.nettoala).toLocaleString('fi')} m²` : '—'],
    ['Valmistumisvuosi', row.valmistumisvuosi || '—'],
    ['Lämmitys', row.lammitys || '—'],
    ['Ilmanvaihto', row.ilmanvaihto || '—'],
    ['Rakennustunnus', row.rakennustunnus || '—'],
  ];

  grid.innerHTML = fields.map(([label, value]) => `
    <div class="bg-slate-50 rounded-lg p-3">
      <div class="text-xs font-semibold text-slate-400 uppercase tracking-wide mb-1">${label}</div>
      <div class="text-sm text-slate-800">${value}</div>
    </div>
  `).join('');

  // Reset omistajapaneeli
  document.getElementById('prh-result').classList.add('hidden');
  document.getElementById('omistaja-error').classList.add('hidden');
  document.getElementById('ai-loading').classList.add('hidden');
  document.getElementById('ai-result').classList.add('hidden');
  document.getElementById('ai-result').innerHTML = '';
  document.getElementById('ai-haku-btn').disabled = false;

  // ── Nimi-fi: näytetään suoraan jos sisältää AS OY / KOY ──
  const nimiBlock = document.getElementById('nimi-ehdotus-block');
  const nimi = row.nimi || '';
  const onYhtio = /as\.?\s*oy|asunto\s*oy|kiinteist[oö][\s.]*oy|koy/i.test(nimi);
  const tallennettu = state.tallennetut[row.rakennustunnus];

  if (nimi) {
    // Ehdotettu nimi: tallennettu > siivottu yhtiönimi > alkuperäinen nimi
    const siivottu = nimi.replace(/^[^,]+,\s*/i, m =>
      /as\.?\s*oy|asunto\s*oy|kiinteist/i.test(m) ? m : ''
    ).trim();
    const nimiEhdotus = tallennettu?.yhtio_nimi || siivottu || nimi;

    nimiBlock.innerHTML = `
      <div class="bg-slate-50 border border-slate-200 rounded-lg p-3">
        <div class="text-xs font-semibold text-slate-500 uppercase tracking-wide mb-2">Yhtiönimi</div>
        <div class="flex items-center gap-2">
          <input id="nimi-muokkaus-input" type="text" value="${nimiEhdotus.replace(/"/g,'&quot;')}"
            class="flex-1 text-sm font-medium border border-slate-200 rounded-lg px-3 py-1.5 focus:ring-2 focus:ring-slate-400 focus:outline-none bg-white">
          <button onclick="tallennaMuokattuNimi()"
            class="flex-shrink-0 text-xs bg-slate-700 text-white px-3 py-1.5 rounded-lg hover:bg-slate-800 transition-colors font-medium whitespace-nowrap">
            Tallenna
          </button>
        </div>
        ${tallennettu?.yhtio_nimi
          ? `<div class="text-xs text-emerald-600 mt-1.5">✓ Tallennettu${tallennettu.ytunnus ? ' · Y-tunnus: ' + tallennettu.ytunnus : ''}</div>`
          : `<div class="text-xs text-slate-400 mt-1.5">Muokkaa tarvittaessa ja tallenna</div>`
        }
      </div>`;
  } else {
    nimiBlock.innerHTML = '';
  }

  // Täytä omistajahakukenttä: käytetään nimi-fi:tä hakusanana omistajalle
  const osoite  = row.osoite || '';
  const yhtioNimi = onYhtio
    ? nimi.replace(/^[^,]+,\s*/i, m => /as\.?\s*oy|asunto\s*oy|kiinteist/i.test(m) ? m : '').trim()
    : nimi;
  // Näytä tallennettu omistaja tai piilota
  const omistajaNykyinen = document.getElementById('omistaja-nykyinen');
  const omistajaNykyinenNimi = document.getElementById('omistaja-nykyinen-nimi');
  const savedOmistaja = tallennettu?.omistaja || '';
  if (savedOmistaja) {
    omistajaNykyinenNimi.textContent = savedOmistaja;
    omistajaNykyinen.classList.remove('hidden');
  } else {
    omistajaNykyinen.classList.add('hidden');
  }

  // Täytä manuaalikenttä tallennetulla omistajalla jos on
  document.getElementById('omistaja-manuaali').value = savedOmistaja;

  // Kontaktoitu-nappi
  const kontaktoituAt = tallennettu?.kontaktoitu_at || '';
  const kBtn = document.getElementById('kontaktoitu-btn');
  const kPvm = document.getElementById('kontaktoitu-pvm');
  if (kontaktoituAt) {
    kBtn.textContent = 'Merkitse ei-kontaktoiduksi';
    kBtn.className = 'text-sm font-medium px-4 py-2 rounded-lg border transition-colors bg-emerald-50 border-emerald-300 text-emerald-700 hover:bg-emerald-100';
    kPvm.textContent = 'Kontaktoitu: ' + kontaktoituAt;
  } else {
    kBtn.textContent = 'Merkitse kontaktoiduksi';
    kBtn.className = 'text-sm font-medium px-4 py-2 rounded-lg border transition-colors bg-white border-slate-200 text-slate-600 hover:bg-slate-50';
    kPvm.textContent = '';
  }

  // Muistiinpanot
  document.getElementById('muistiinpanot-input').value = tallennettu?.muistiinpanot || '';

  document.getElementById('modal').classList.add('open');
}

async function toggleKontaktoitu() {
  if (!currentRow) return;
  const rakennustunnus = currentRow.rakennustunnus;
  const tallennettu = state.tallennetut[rakennustunnus];
  const onKontaktoitu = !!(tallennettu?.kontaktoitu_at);

  const res = await fetch('/api/merkitse-kontaktoiduksi', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({rakennustunnus, poista: onKontaktoitu}),
  });
  const data = await res.json();
  if (data.success) {
    if (!state.tallennetut[rakennustunnus]) state.tallennetut[rakennustunnus] = {rakennustunnus};
    state.tallennetut[rakennustunnus].kontaktoitu_at = data.kontaktoitu_at || '';
    // Päivitä nappi
    const kBtn = document.getElementById('kontaktoitu-btn');
    const kPvm = document.getElementById('kontaktoitu-pvm');
    if (data.kontaktoitu_at) {
      kBtn.textContent = 'Merkitse ei-kontaktoiduksi';
      kBtn.className = 'text-sm font-medium px-4 py-2 rounded-lg border transition-colors bg-emerald-50 border-emerald-300 text-emerald-700 hover:bg-emerald-100';
      kPvm.textContent = 'Kontaktoitu: ' + data.kontaktoitu_at;
    } else {
      kBtn.textContent = 'Merkitse kontaktoiduksi';
      kBtn.className = 'text-sm font-medium px-4 py-2 rounded-lg border transition-colors bg-white border-slate-200 text-slate-600 hover:bg-slate-50';
      kPvm.textContent = '';
    }
    // Päivitä kortti
    const card = document.querySelector(`[data-tunnus="${rakennustunnus}"]`);
    if (card) {
      const chipsDiv = card.querySelector('.flex.gap-1\\.5');
      const existingChip = chipsDiv?.querySelector('.kontaktoitu-chip');
      if (data.kontaktoitu_at && chipsDiv && !existingChip) {
        const chip = document.createElement('span');
        chip.className = 'kontaktoitu-chip text-xs bg-emerald-100 text-emerald-700 px-2 py-0.5 rounded-full font-medium';
        chip.textContent = '✓ Kontaktoitu';
        chipsDiv.appendChild(chip);
      } else if (!data.kontaktoitu_at && existingChip) {
        existingChip.remove();
      }
    }
  }
}

async function tallennaMuistiinpano() {
  if (!currentRow) return;
  const teksti = document.getElementById('muistiinpanot-input').value.trim();
  const rakennustunnus = currentRow.rakennustunnus;
  const res = await fetch('/api/tallenna-muistiinpano', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({rakennustunnus, teksti}),
  });
  const data = await res.json();
  if (data.success) {
    if (!state.tallennetut[rakennustunnus]) state.tallennetut[rakennustunnus] = {rakennustunnus};
    state.tallennetut[rakennustunnus].muistiinpanot = teksti;
    const btn = document.querySelector('button[onclick="tallennaMuistiinpano()"]');
    if (btn) { btn.textContent = '✓ Tallennettu'; setTimeout(() => btn.textContent = 'Tallenna muistiinpano', 1500); }
  }
}

async function kaynnistaMassahaku() {
  const statusEl = document.getElementById('massahaku-status');
  statusEl.textContent = 'Haetaan näkyvät rakennukset...';
  statusEl.classList.remove('hidden');

  // Kerää nykyiset kortit
  const cards = document.querySelectorAll('#table-body [data-tunnus]');
  const rakennukset = [];
  cards.forEach(card => {
    if (card._rowData) rakennukset.push(card._rowData);
  });

  if (!rakennukset.length) {
    statusEl.textContent = 'Ei rakennuksia — hae ensin tuloksia';
    return;
  }

  const erä = rakennukset.slice(0, 20);
  statusEl.textContent = `AI hakee omistajia ${erä.length} rakennukselle...`;

  const res = await fetch('/api/massahaku', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({rakennukset: erä}),
  });
  const data = await res.json();
  if (!data.success) { statusEl.textContent = '⚠️ ' + data.error; return; }

  let tallennettu = 0, ohitettu = 0;
  data.tulokset.forEach(t => {
    if (t.skip) { ohitettu++; return; }
    if (t.tallennettu && t.omistaja) {
      tallennettu++;
      if (!state.tallennetut[t.rakennustunnus]) state.tallennetut[t.rakennustunnus] = {rakennustunnus: t.rakennustunnus};
      state.tallennetut[t.rakennustunnus].omistaja = t.omistaja;
      paivitaOmistajaChip(t.rakennustunnus, t.omistaja);
    }
  });

  statusEl.textContent = `✓ Valmis — ${tallennettu} omistajaa tallennettu, ${ohitettu} ohitettu (jo tallennettu)`;
  setTimeout(() => statusEl.classList.add('hidden'), 5000);
}

function closeModal() {
  document.getElementById('modal').classList.remove('open');
  if (state.selectedRow) state.selectedRow.classList.remove('selected');
  state.selectedRow = null;
}

document.getElementById('modal').addEventListener('click', e => {
  if (e.target === document.getElementById('modal')) closeModal();
});

// ── TALLENNA OMISTAJA ──
async function tallennaOmistaja(omistajaNimi) {
  if (!omistajaNimi || !omistajaNimi.trim()) return;
  if (!currentRow) return;
  const rakennustunnus = currentRow.rakennustunnus;

  const res = await fetch('/api/hyvaksy-yhtio', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({
      rakennustunnus,
      osoite:    currentRow.osoite,
      yhtio_nimi: currentRow.nimi || '',   // yhtiönimi CSV:stä
      omistaja:  omistajaNimi.trim(),
    }),
  });
  const data = await res.json();
  if (data.success) {
    // Päivitä state
    if (!state.tallennetut[rakennustunnus])
      state.tallennetut[rakennustunnus] = {rakennustunnus};
    state.tallennetut[rakennustunnus].omistaja  = omistajaNimi.trim();
    state.tallennetut[rakennustunnus].yhtio_nimi = data.yhtio_nimi || currentRow.nimi || '';

    // Näytä vahvistus
    document.getElementById('prh-lista').innerHTML = `
      <div class="bg-emerald-50 border border-emerald-200 rounded-lg p-3 flex items-center gap-3">
        <svg class="w-5 h-5 text-emerald-600 flex-shrink-0" fill="none" stroke="currentColor" viewBox="0 0 24 24">
          <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M5 13l4 4L19 7"/>
        </svg>
        <div>
          <div class="font-semibold text-emerald-800">${omistajaNimi}</div>
          <div class="text-xs text-emerald-600 mt-0.5">Tallennettu omistaja</div>
        </div>
      </div>`;
    document.getElementById('prh-result').classList.remove('hidden');

    // Päivitä omistaja-manuaalikenttä
    document.getElementById('omistaja-manuaali').value = omistajaNimi.trim();

    // Päivitä "omistaja-nykyinen" -osio
    const nykyinenDiv = document.getElementById('omistaja-nykyinen');
    document.getElementById('omistaja-nykyinen-nimi').textContent = omistajaNimi.trim();
    nykyinenDiv.classList.remove('hidden');

    // Päivitä kortti listassa
    paivitaOmistajaChip(rakennustunnus, omistajaNimi.trim());
  } else {
    alert('Tallennus epäonnistui: ' + data.error);
  }
}

async function tallennaManuaaliOmistaja() {
  const v = document.getElementById('omistaja-manuaali').value.trim();
  if (!v) { document.getElementById('omistaja-manuaali').focus(); return; }
  await tallennaOmistaja(v);
}

function paivitaOmistajaChip(rakennustunnus, omistajaNimi) {
  const card = document.querySelector(`[data-tunnus="${rakennustunnus}"]`);
  if (!card) return;
  const chipsDiv = card.querySelector('.flex.gap-1\\.5');
  if (!chipsDiv) {
    // Chips-rivi puuttuu kokonaan — luodaan se
    const contentDiv = card.querySelector('.flex-1');
    if (contentDiv) {
      const newChips = document.createElement('div');
      newChips.className = 'flex gap-1.5 mt-2 flex-wrap';
      if (omistajaNimi) {
        newChips.innerHTML = `<span class="omistaja-chip text-xs bg-violet-100 text-violet-700 px-2 py-0.5 rounded-full font-medium">${omistajaNimi}</span>`;
      }
      contentDiv.appendChild(newChips);
    }
    return;
  }
  // Päivitä tai poista olemassa oleva omistaja-chip
  const existingChip = chipsDiv.querySelector('.omistaja-chip');
  if (omistajaNimi) {
    if (existingChip) {
      existingChip.textContent = omistajaNimi;
    } else {
      const chip = document.createElement('span');
      chip.className = 'omistaja-chip text-xs bg-violet-100 text-violet-700 px-2 py-0.5 rounded-full font-medium';
      chip.textContent = omistajaNimi;
      chipsDiv.appendChild(chip);
    }
  } else {
    if (existingChip) existingChip.remove();
  }
}

function korvaaOmistaja() {
  // Piilota nykyinen-näkymä ja fokusoi manuaalikenttä
  document.getElementById('omistaja-nykyinen').classList.add('hidden');
  const input = document.getElementById('omistaja-manuaali');
  input.value = '';
  input.focus();
}

async function tyhjennaOmistaja() {
  if (!currentRow) return;
  if (!confirm('Poistetaanko omistajatieto?')) return;
  const rakennustunnus = currentRow.rakennustunnus;

  const res = await fetch('/api/hyvaksy-yhtio', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({
      rakennustunnus,
      osoite:     currentRow.osoite,
      yhtio_nimi: currentRow.nimi || '',
      omistaja:   '',
      tyhjenna_omistaja: true,
    }),
  });
  const data = await res.json();
  if (data.success) {
    if (state.tallennetut[rakennustunnus]) {
      state.tallennetut[rakennustunnus].omistaja = '';
    }
    document.getElementById('omistaja-nykyinen').classList.add('hidden');
    document.getElementById('omistaja-manuaali').value = '';

    // Päivitä kortti listassa
    paivitaOmistajaChip(rakennustunnus, '');
  } else {
    alert('Virhe: ' + data.error);
  }
}

// Säilytetään hyvaksyYhtio yhtiönimen tallentamiseen (nimi-ehdotus-block)
async function tallennaMuokattuNimi() {
  const input = document.getElementById('nimi-muokkaus-input');
  if (!input) return;
  const v = input.value.trim();
  if (!v) { input.focus(); return; }
  await hyvaksyYhtio(v, '');
  // Päivitä huomioteksti
  const nota = input.parentElement?.nextElementSibling;
  if (nota) nota.innerHTML = `<div class="text-xs text-emerald-600 mt-1.5">✓ Tallennettu</div>`;
}

async function hyvaksyYhtio(yhtioNimi, ytunnus) {
  if (!currentRow) return;
  const rakennustunnus = currentRow.rakennustunnus;
  if (!rakennustunnus) { alert('Rakennustunnusta ei löydy'); return; }

  const res = await fetch('/api/hyvaksy-yhtio', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({rakennustunnus, osoite: currentRow.osoite, yhtio_nimi: yhtioNimi, ytunnus}),
  });
  const data = await res.json();
  if (data.success) {
    if (!state.tallennetut[rakennustunnus])
      state.tallennetut[rakennustunnus] = {rakennustunnus};
    state.tallennetut[rakennustunnus].yhtio_nimi = yhtioNimi;

    // Päivitä nimi-block → näytä vihreä
    document.getElementById('nimi-ehdotus-block').innerHTML = `
      <div class="bg-emerald-50 border border-emerald-200 rounded-lg p-3 flex items-center gap-3">
        <svg class="w-5 h-5 text-emerald-600 flex-shrink-0" fill="none" stroke="currentColor" viewBox="0 0 24 24">
          <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M5 13l4 4L19 7"/>
        </svg>
        <div>
          <div class="font-semibold text-emerald-800">${yhtioNimi}</div>
          <div class="text-xs text-emerald-600">Tallennettu yhtiönimi</div>
        </div>
      </div>`;
  } else {
    alert('Tallennus epäonnistui: ' + data.error);
  }
}

// ── AI-OMISTAJAHAKU ──
async function haeOmistajaAI() {
  if (!currentRow) return;

  const btn = document.getElementById('ai-haku-btn');
  const loading = document.getElementById('ai-loading');
  const resultEl = document.getElementById('ai-result');

  btn.disabled = true;
  loading.classList.remove('hidden');
  resultEl.classList.add('hidden');
  resultEl.innerHTML = '';

  try {
    const res = await fetch('/api/omistaja-ai', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({rakennus: currentRow}),
    });
    const data = await res.json();
    loading.classList.add('hidden');
    btn.disabled = false;

    if (!data.success) {
      resultEl.innerHTML = `<div class="text-xs text-red-600 bg-red-50 rounded-lg p-3 whitespace-pre-wrap">${data.error}</div>`;
      resultEl.classList.remove('hidden');
      return;
    }

    const r = data.result;
    const owner   = r.owner   || {};
    const contact = r.contact || {};
    const confColor = owner.confidence === 'HIGH' ? 'emerald' : owner.confidence === 'MEDIUM' ? 'amber' : 'orange';
    const confLabel = owner.confidence === 'HIGH' ? 'Varma' : owner.confidence === 'MEDIUM' ? 'Todennäköinen' : 'Epävarma';
    const mqColor = contact.match_quality === 'DIRECT' ? 'emerald' : contact.match_quality === 'CLOSE' ? 'amber' : 'slate';
    const mqLabel = contact.match_quality === 'DIRECT' ? 'Suora' : contact.match_quality === 'CLOSE' ? 'Läheinen' : 'Portfolio';

    const evidenceHtml = (r.evidence || []).map(e => `
      <li class="flex gap-1.5">
        <span class="text-xs mt-0.5 text-slate-400">·</span>
        <span class="text-xs text-slate-600">${typeof e === 'string' ? e : e.description || ''}</span>
      </li>`).join('');

    const contactHtml = contact.name ? `
      <div class="border border-slate-100 rounded-lg p-3 space-y-1 bg-slate-50">
        <div class="flex items-center justify-between gap-2">
          <div class="text-xs font-semibold text-slate-500 uppercase tracking-wide">Yhteyshenkilö</div>
          ${contact.match_quality ? `<span class="text-xs bg-${mqColor}-100 text-${mqColor}-700 px-2 py-0.5 rounded-full">${mqLabel}</span>` : ''}
        </div>
        <div class="font-semibold text-slate-800 text-sm">${contact.name}</div>
        ${contact.role    ? `<div class="text-xs text-slate-500">${contact.role}${contact.company ? ' · ' + contact.company : ''}</div>` : ''}
        ${contact.email   ? `<div class="text-xs text-blue-600 mt-1"><a href="mailto:${contact.email}">${contact.email}</a></div>` : ''}
        ${contact.phone   ? `<div class="text-xs text-slate-600">${contact.phone}</div>` : ''}
      </div>` : '';

    resultEl.innerHTML = `
      <div class="bg-white border border-violet-100 rounded-xl p-4 space-y-3">

        <!-- Omistaja -->
        <div class="flex items-start justify-between gap-3">
          <div class="min-w-0">
            <div class="text-xs font-semibold text-slate-400 uppercase tracking-wide mb-0.5">Omistaja</div>
            <div class="font-bold text-slate-800 text-base">${owner.name || '—'}</div>
            ${r.asset_manager ? `<div class="text-xs text-slate-500 mt-0.5">Asset manager: ${r.asset_manager}</div>` : ''}
            ${r.recognized_asset ? `<div class="text-xs text-violet-600 mt-0.5">✓ Tunnistettu kiinteistö</div>` : ''}
          </div>
          <div class="flex flex-col items-end gap-2 flex-shrink-0">
            <span class="text-xs font-semibold bg-${confColor}-100 text-${confColor}-700 px-2 py-0.5 rounded-full">${confLabel}</span>
            ${owner.name ? `<button onclick="tallennaOmistaja('${owner.name.replace(/'/g,"\\'")}')"
              class="text-xs bg-emerald-600 text-white px-3 py-1.5 rounded-lg hover:bg-emerald-700 font-medium whitespace-nowrap">
              Tallenna omistaja
            </button>` : ''}
          </div>
        </div>

        <!-- Yhteyshenkilö -->
        ${contactHtml}

        <!-- Evidence -->
        ${evidenceHtml ? `<ul class="space-y-1 border-t border-slate-100 pt-2">${evidenceHtml}</ul>` : ''}

        <!-- Perustelu -->
        ${r.reasoning ? `<div class="text-xs text-slate-500 italic border-t border-slate-100 pt-2">${r.reasoning}</div>` : ''}
      </div>`;
    resultEl.classList.remove('hidden');

  } catch(err) {
    loading.classList.add('hidden');
    btn.disabled = false;
    resultEl.innerHTML = `<div class="text-xs text-red-600 bg-red-50 rounded-lg p-3">Verkkovirhe: ${err.message}</div>`;
    resultEl.classList.remove('hidden');
  }
}


init();
</script>
</body>
</html>"""


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    print(f"🌐  Avaa selain: http://localhost:{port}\n")
    app.run(debug=False, port=port, host='0.0.0.0')
